from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
# import pandas as pd # Removed for size optimization
import csv
import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

import requests
from bs4 import BeautifulSoup
# from duckduckgo_search import DDGS # Moved to local scope for safety
import time
import re
from urllib.parse import urlparse

import json
import os
import redis
from dotenv import load_dotenv

# Load environment variables from .env file (for local dev)
load_dotenv()

# Redis / Vercel KV Configuration
KV_URL = os.environ.get("KV_URL") or os.environ.get("REDIS_URL")
redis_client = None

if KV_URL:
    try:
        redis_client = redis.from_url(KV_URL)
        redis_client.ping()
        print("Connected to Vercel KV (Redis)")
    except Exception as e:
        print(f"Failed to connect to Redis: {e}")
        redis_client = None

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)

from ai_classifier import analyze_with_groq

# --- Configuration ---
CORRECTIONS_FILE = "corrections.json"
USER_CORRECTIONS = {}

def load_corrections():
    global USER_CORRECTIONS
    
    # 1. Try Redis first
    if redis_client:
        try:
            # HGETALL returns byte keys/values, need to decode
            data = redis_client.hgetall("corrections")
            USER_CORRECTIONS = {k.decode('utf-8'): v.decode('utf-8') for k, v in data.items()}
            return
        except Exception as e:
            print(f"Redis Load Error: {e}")
            # Fallback to local file if Redis fails
            
    # 2. Fallback to Local File
    if os.path.exists(CORRECTIONS_FILE):
        try:
            with open(CORRECTIONS_FILE, 'r', encoding='utf-8') as f:
                USER_CORRECTIONS = json.load(f)
        except:
            USER_CORRECTIONS = {}

def normalize_key(name):
    # Centralized normalization for consistent key generation
    if not name: return ""
    return name.upper().strip()

def save_correction(name, sector):
    global USER_CORRECTIONS
    # Normalize key: uppercase without spaces/special chars for robust matching
    key = normalize_key(name)
    USER_CORRECTIONS[key] = sector
    
    # 1. Save to Redis
    if redis_client:
        try:
            redis_client.hset("corrections", key, sector)
        except Exception as e:
            print(f"Redis Save Error: {e}")
            
    # 2. Save to Local File (Always try to keep in sync if possible, or for dev)
    try:
        with open(CORRECTIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(USER_CORRECTIONS, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving correction: {e}")

# Load on startup
load_corrections()

# --- Custom Sectors Configuration ---
CUSTOM_SECTORS_FILE = "custom_sectors.json"
CUSTOM_SECTORS = []

def load_custom_sectors():
    global CUSTOM_SECTORS
    
    # 1. Try Redis
    if redis_client:
        try:
            # Storing as a single JSON string for simplicity (or could be a Redis Set)
            data = redis_client.get("custom_sectors")
            if data:
                CUSTOM_SECTORS = json.loads(data)
                return
        except Exception as e:
             print(f"Redis Custom Sectors Load Error: {e}")
             
    # 2. Fallback Local
    if os.path.exists(CUSTOM_SECTORS_FILE):
        try:
            with open(CUSTOM_SECTORS_FILE, 'r', encoding='utf-8') as f:
                CUSTOM_SECTORS = json.load(f)
        except:
            CUSTOM_SECTORS = []

def save_custom_sectors():
    
    # 1. Redis
    if redis_client:
        try:
            redis_client.set("custom_sectors", json.dumps(CUSTOM_SECTORS))
        except Exception as e:
             print(f"Redis Custom Sectors Save Error: {e}")

    # 2. Local
    try:
        with open(CUSTOM_SECTORS_FILE, 'w', encoding='utf-8') as f:
            json.dump(CUSTOM_SECTORS, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving custom sectors: {e}")

load_custom_sectors()

# --- Comprehensive Region Mapping (The Trap Anti-Fail) ---
# --- Comprehensive Region Mapping (The Trap Anti-Fail) ---
DEPT_TO_REGION = {
    "01": "Auvergne-Rhône-Alpes", "02": "Hauts-de-France", "03": "Auvergne-Rhône-Alpes", "04": "Provence-Alpes-Côte d'Azur", "05": "Provence-Alpes-Côte d'Azur",
    "06": "Provence-Alpes-Côte d'Azur", "07": "Auvergne-Rhône-Alpes", "08": "Grand Est", "09": "Occitanie", "10": "Grand Est",
    "11": "Occitanie", "12": "Occitanie", "13": "Provence-Alpes-Côte d'Azur", "14": "Normandie", "15": "Auvergne-Rhône-Alpes",
    "16": "Nouvelle-Aquitaine", "17": "Nouvelle-Aquitaine", "18": "Centre-Val de Loire", "19": "Nouvelle-Aquitaine", "21": "Bourgogne-Franche-Comté",
    "22": "Bretagne", "23": "Nouvelle-Aquitaine", "24": "Nouvelle-Aquitaine", "25": "Bourgogne-Franche-Comté", "26": "Auvergne-Rhône-Alpes",
    "27": "Normandie", "28": "Centre-Val de Loire", "29": "Bretagne", "2A": "Corse", "2B": "Corse",
    "30": "Occitanie", "31": "Occitanie", "32": "Occitanie", "33": "Nouvelle-Aquitaine", "34": "Occitanie",
    "35": "Bretagne", "36": "Centre-Val de Loire", "37": "Centre-Val de Loire", "38": "Auvergne-Rhône-Alpes", "39": "Bourgogne-Franche-Comté",
    "40": "Nouvelle-Aquitaine", "41": "Centre-Val de Loire", "42": "Auvergne-Rhône-Alpes", "43": "Auvergne-Rhône-Alpes", "44": "Pays de la Loire",
    "45": "Centre-Val de Loire", "46": "Occitanie", "47": "Nouvelle-Aquitaine", "48": "Occitanie", "49": "Pays de la Loire",
    "50": "Normandie", "51": "Grand Est", "52": "Grand Est", "53": "Pays de la Loire", "54": "Grand Est",
    "55": "Grand Est", "56": "Bretagne", "57": "Grand Est", "58": "Bourgogne-Franche-Comté", "59": "Hauts-de-France",
    "60": "Hauts-de-France", "61": "Normandie", "62": "Hauts-de-France", "63": "Auvergne-Rhône-Alpes", "64": "Nouvelle-Aquitaine",
    "65": "Occitanie", "66": "Occitanie", "67": "Grand Est", "68": "Grand Est", "69": "Auvergne-Rhône-Alpes",
    "70": "Bourgogne-Franche-Comté", "71": "Bourgogne-Franche-Comté", "72": "Pays de la Loire", "73": "Auvergne-Rhône-Alpes", "74": "Auvergne-Rhône-Alpes",
    "75": "Île-de-France", "76": "Normandie", "77": "Île-de-France", "78": "Île-de-France", "79": "Nouvelle-Aquitaine",
    "80": "Hauts-de-France", "81": "Occitanie", "82": "Occitanie", "83": "Provence-Alpes-Côte d'Azur", "84": "Provence-Alpes-Côte d'Azur",
    "85": "Pays de la Loire", "86": "Nouvelle-Aquitaine", "87": "Nouvelle-Aquitaine", "88": "Grand Est", "89": "Bourgogne-Franche-Comté",
    "90": "Bourgogne-Franche-Comté", "91": "Île-de-France", "92": "Île-de-France", "93": "Île-de-France", "94": "Île-de-France",
    "95": "Île-de-France", "971": "Guadeloupe", "972": "Martinique", "973": "Guyane", "974": "La Réunion", "976": "Mayotte"
}

SECTOR_CONFIG = {
    "Agriculture / Livestock / Seafood": {
        "naf_prefixes": ["01", "02", "03"],
        "keywords": ["agriculture", "élevage", "pêche", "agricole", "ferme", "bio", "tracteur", "champs", "vigne", "viticulture", "horticulture", "maraichage", "bétail", "aquaculture", "farming", "livestock", "seafood", "crops", "agri", "agro", "vignoble", "éleveur", "céréales", "semences", "fisheries"]
    },
    "Banking": {
        "naf_prefixes": ["641"],
        "keywords": ["banque", "crédit", "bancaire", "compte", "livret", "cb", "bank", "banking", "loan", "credit", "bnp", "société générale", "crédit agricole", "bpce", "lender", "mortgage", "prêt", "emprunt", "financement", "épargne"]
    },
    "Chemicals": {
        "naf_prefixes": ["20"],
        "keywords": ["chimie", "laboratoire", "molécules", "réactif", "polymère", "plastique", "chimique", "petrochemical", "chemicals", "chemistry", "lab", "solvay", "arkema", "air liquide", "gas", "gaz", "azote", "hydrogène", "composites", "resins", "paint", "peinture", "coatings"]
    },
    "Communication / Media & Entertainment / Telecom": {
        "naf_prefixes": ["59", "60", "61", "63"],
        "keywords": ["télécom", "média", "publicité", "fibre", "internet", "presse", "journal", "tv", "radio", "marketing", "agence", "communication", "entertainment", "telecom", "broadcasting", "advertising", "media", "orange", "sfr", "bouygues", "free", "publicis", "havas", "digital agency", "rédaction", "news", "contenu", "publishing", "édition", "production audiovisuelle", "streaming"]
    },
    "Construction": {
        "naf_prefixes": ["41", "42", "43"],
        "keywords": ["btp", "construction", "bâtiment", "génie civil", "infrastructure", "travaux", "architecture", "maçonnerie", "électicité", "plomberie", "architect", "builder", "contractor", "civil", "renovation", "vinci", "eiffage", "bouygues construction", "rénovation", "agencement", "menuiserie", "charpente", "promoteur", "immobilier neuf", "engineering", "ingénierie bâtiment", "hvac", "étanchéité"]
    },
    "Consulting / IT Services": {
        "naf_prefixes": ["62", "631", "582", "702", "692", "7112", "712", "732", "74"],
        "keywords": ["conseil", "consulting", "esn", "stratégie", "audit", "expertise", "ingénierie", "rub", "management", "digital", "transformation", "it services", "système d'information", "data", "advisory", "capgemini", "deloitte", "kpmg", "pwc", "mckinsey", "bain", "bcg", "accenture", "sogeti", "sopra", "wavestone", "alteca", "umanis", "devops", "cloud computing", "cybersécurité", "business intelligence", "big data", "agile", "scrum", "moa", "moe", "change management"]
    },
    "CPG (Consumer Packaged Goods)": {
        "naf_prefixes": ["204"],
        "keywords": ["fmcg", "biens de consommation", "hygiène", "produits ménagers", "cosmétique", "beauté", "parfum", "shampoing", "savon", "lessive", "cpg", "consumer goods", "l'oréal", "procter", "gamble", "unilever", "danone", "nestlé", "henkel", "persil", "dash", "ariel", "schwarzkopf", "nivea", "dove", "maquillage", "makeup", "skincare", "soin", "personal care", "toiletries"]
    },
    "Education": {
        "naf_prefixes": ["85"],
        "keywords": ["éducation", "formation", "école", "université", "training", "learning", "elearning", "edtech", "campus", "formation continue", "school", "university", "academy", "college", "enseignement", "pédagogie", "cours", "tutoring", "soutien scolaire", "mba", "master", "licence", "certification"]
    },
    "Energy / Utilities": {
        "naf_prefixes": ["35", "36", "37", "38", "39"],
        "keywords": ["énergie", "électricité", "gaz", "eau", "déchets", "environnement", "recyclage", "solaire", "éolien", "nucléaire", "oil", "petrol", "renewables", "green", "carbon", "hydrogen", "edf", "engie", "total", "veolia", "suez", "photovoltaïque", "biomasse", "hydro", "grid", "réseau électrique", "assainissement", "waste management", "energy", "batteries", "charging"]
    },
    "Finance / Real Estate": {
        "naf_prefixes": ["64", "66", "68"],
        "keywords": ["finance", "financial", "services financiers", "immobilier", "investissement", "gestion d'actifs", "courtier", "syndic", "promoteur", "real estate", "realty", "property", "logement", "immo", "wealth", "fintech", "payment", "trading", "crypto", "blockchain", "vc", "private equity", "fund", "foncia", "nexity", "asset management", "patrimoine", "défiscalisation", "location", "vente immobilière", "agency", "investor", "capital", "holding", "reit"]
    },
    "Food / Beverages": {
        "naf_prefixes": ["10", "11"],
        "keywords": ["agroalimentaire", "aliments", "boissons", "food", "beverage", "vin", "spiritueux", "bière", "champagne", "nutrition", "snack", "dairy", "laitier", "viande", "boulangerie", "traiteur", "épicerie", "confiserie", "chocolat", "surgelés", "frozen", "drinks", "juice", "jus", "distillery", "brewery", "winery", "bio food", "organic", "restaurant supply"]
    },
    "Healthcare / Medical Services": {
        "naf_prefixes": ["86", "87", "88"],
        "keywords": ["santé", "clinique", "hôpital", "soins", "médecin", "infirmier", "ehpad", "médical", "chirurgie", "patient", "healthcare", "medical", "hospital", "clinic", "care", "doctor", "diagnostic", "radiologie", "dentaire", "kine", "ramsay", "elsan", "korian", "orpea", "nursing", "home care", "aide à domicile", "analyse", "labo", "biologie", "medtech", "e-health"]
    },
    "Hotels / Restaurants": {
        "naf_prefixes": ["55", "56"],
        "keywords": ["hôtel", "restaurant", "tourisme", "hébergement", "camping", "voyage", "bar", "café", "brasserie", "cuisine", "hotel", "hospitality", "tourism", "restaurant", "catering", "accor", "club med", "sodexo", "elior", "travel", "resort", "vacances", "booking", "chef", "gastronomie", "food service", "fast food"]
    },
    "Insurance / Mutual Health Insurance": {
        "naf_prefixes": ["65"],
        "keywords": ["assurance", "mutuelle", "courtage", "assureur", "prévoyance", "risques", "insurance", "underwriting", "axa", "allianz", "generali", "maif", "macif", "groupama", "malakoff", "ag2r", "harmonie", "protection sociale", "sinistre", "broker", "reinsurance", "réassurance", "insurtech"]
    },
    "Luxury": {
        "naf_prefixes": ["141", "142", "151", "152"],
        "keywords": ["luxe", "prestige", "haute couture", "joaillerie", "maroquinerie", "palace", "luxury", "fashion", "jewelry", "premium", "high-end", "mode", "vêtement", "chaussures", "shoes", "wear", "apparel", "lvmh", "kering", "hermès", "chanel", "dior", "vuitton", "gucci", "prada", "rolex", "cartier", "bijoux", "diamant", "montres", "watches", "perfumery"]
    },
    "Manufacturing / Industry": {
        "naf_prefixes": ["13", "14", "15", "16", "17", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31", "32", "33"],
        "keywords": ["industrie", "usine", "fabrication", "mécanique", "métallurgie", "plasturgie", "assemblage", "production", "machine", "outil", "industriel", "manufacturing", "industry", "factory", "plant", "metal", "machinery", "automotive", "aéronautique", "aerospace", "defense", "textile", "imprimerie", "packaging", "saint-gobain", "schneider", "legrand", "michelin", "stellantis", "renault", "airbus", "thales", "safran", "dassault", "alstom", "composants", "robotics", "automation", "electronics assembly"]
    },
    "Not For Profit": {
        "naf_prefixes": ["94", "91"],
        "keywords": ["association", "fondation", "ong", "non-profit", "charity", "bénévole", "social", "humanitaire", "syndicat", "union", "club", "croix rouge", "secours populaire", "médecins sans frontières", "unicef", "caritas", "aide", "solidarité", "non lucratif", "philanthropy"]
    },
    "Pharmaceutics": {
        "naf_prefixes": ["21"],
        "keywords": ["pharmacie", "médicament", "biotech", "laboratoire", "vaccin", "recherche", "molécule", "thérapie", "pharmaceutical", "pharma", "drug", "biotechnology", "medicine", "lifescience", "sanofi", "servier", "pfizer", "moderna", "astrazeneca", "bayer", "novartis", "roche", "lilly", "clinical trials", "essais cliniques", "cro"]
    },
    "Public administration & government": {
        "naf_prefixes": ["84"],
        "keywords": ["mairie", "préfecture", "ministère", "collectivité", "public", "etat", "government", "administration", "caisse", "caf", "urssaf", "pole emploi", "france travail", "ambassade", "consulat", "département", "région", "agglomération", "commune", "service public"]
    },
    "Retail": {
        "naf_prefixes": ["45", "46", "47"],
        "keywords": ["commerce", "vente", "magasin", "boutique", "supermarché", "distribution", "retail", "store", "shop", "e-commerce", "marketplace", "grossiste", "grand magasin", "shopping", "mall", "outlet", "franchise", "carrefour", "auchan", "leclerc", "decathlon", "fnac", "darty", "amazon", "cdiscount", "bricolage", "jardinage", "ameublement", "fashion retail", "grocery", "point de vente", "wholesaler"]
    },
    "HR / Recruitment / Interim": {
        "naf_prefixes": ["78"],
        "keywords": ["intérim", "recrutement", "rh", "ressources humaines", "agence d'emploi", "staffing", "recruitment", "chasseur de tête", "talent", "manpower", "adecco", "randstad", "crit", "synergie", "proman", "michael page", "hays", "robert half", "headhunting", "jobs", "emplois", "carrière", "paye", "payroll"]
    },
    "Tech / Software": {
        "naf_prefixes": ["582", "6201", "6312", "262"],
        "keywords": ["logiciel", "saas", "tech", "software", "application", "ia", "intelligence artificielle", "cloud", "développement", "web", "app", "cybersecurity", "platform", "technology", "developer", "electronics", "hardware", "computer", "start-up", "google", "microsoft", "apple", "meta", "aws", "salesforce", "sap", "oracle", "it", "informatique", "data science", "machine learning", "coding", "programmation", "algorithme", "api", "fintech", "blockchain", "iot", "data center"]
    },
    "Transportation, Logistics & Storage": {
        "naf_prefixes": ["49", "50", "51", "52", "53"],
        "keywords": ["transport", "logistique", "fret", "livraison", "messagerie", "entrepôt", "supply chain", "shipping", "transit", "colis", "airline", "aérien", "avion", "bateau", "compagnie aérienne", "rail", "ferroviaire", "maritime", "port", "sncf", "air france", "maersk", "cma cgm", "dhl", "fedex", "ups", "geodis", "bolloré", "xpo", "container", "cargo", "logistics", "warehouse", "freight"]
    }
}

# --- Competitor Watchlist (Keyrus & Market) ---
COMPETITORS = {
    "ACCENTURE", "CAPGEMINI", "DELOITTE", "PWC", "EY", "KPMG", 
    "SOPRA STERIA", "CGI", "ATOS", "WAVESTONE", "INETUM", 
    "BUSINESS & DECISION", "ARTEFACT", "CONVERTEO", "JEMS", 
    "MICROPOLE", "VISEO", "UMANIS", "DEVOTEAM", "TOLUNA", 
    "BVA", "IPSOS", "KANTAR", "MCKINSEY", 
    "BAIN", "BCG", "BOSTON CONSULTING GROUP", "KEYRUS" # Keeping Keyrus in list but fixing logic? User said "Disney est Keyrus sont toujours des concurrent". 
    # WAIT. User said "Disney est Keyrus sont toujours des concurrent" -> "Disney AND Keyrus remain competitors". 
    # If the user meant "Why are they competitors?", then I should remove them or fix the matching.
    # If the user meant "They ARE competitors" (assertion), then I should default to keeping them if valid, but the user likely reported a bug.
    # The previous turn analysis: "implies an issue... likely that they are incorrectly identified".
    # And my plan said: "Remove KEYRUS from the COMPETITORS list".
    # So I will remove KEYRUS. Line 192 has Keyrus.
    "BVA", "IPSOS", "KANTAR", "MCKINSEY", 
}

NAF_BLACKLIST = ["7010Z", "6420Z"]

# ... (We need to jump to the logic part, but I can't do non-contiguous edits in one block)
# I will do the list revert first, then the logic change in a second step or if the tool supports it... 
# Wait, the prompt says "Revert... Then modify". 
# replace_file_content is for SINGLE CONTIGUOUS BLOCK. 
# I cannot edit line 197 and line 621 in one go.
# I will use multi_replace.

# --- Mappings ---
TRANCHE_EFFECTIFS = {
    "NN": "Non renseigné",
    "00": "0 salarié",
    "01": "1 ou 2 salariés",
    "02": "3 à 5 salariés",
    "03": "6 à 9 salariés",
    "11": "10 à 19 salariés",
    "12": "20 à 49 salariés",
    "21": "50 à 99 salariés",
    "22": "100 à 199 salariés",
    "31": "200 à 249 salariés",
    "32": "250 à 499 salariés",
    "41": "500 à 999 salariés",
    "42": "1 000 à 1 999 salariés",
    "51": "2 000 à 4 999 salariés",
    "52": "5 000 à 9 999 salariés",
    "53": "10 000 salariés et plus"
}

# --- GLOBAL STATIC OVERRIDES (Safety Net) ---
GLOBAL_OVERRIDES = {
    "APPLE": {"Secteur": "Tech / Software", "Nom Officiel": "APPLE INC.", "Adresse": "Cupertino, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "TESLA": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "TESLA INC.", "Adresse": "Austin, TX (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "GOOGLE": {"Secteur": "Tech / Software", "Nom Officiel": "ALPHABET INC.", "Adresse": "Mountain View, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "MICROSOFT": {"Secteur": "Tech / Software", "Nom Officiel": "MICROSOFT CORP", "Adresse": "Redmond, WA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "AMAZON": {"Secteur": "Tech / Software", "Nom Officiel": "AMAZON.COM INC", "Adresse": "Seattle, WA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "META": {"Secteur": "Tech / Software", "Nom Officiel": "META PLATFORMS", "Adresse": "Menlo Park, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "FACEBOOK": {"Secteur": "Tech / Software", "Nom Officiel": "META PLATFORMS", "Adresse": "Menlo Park, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "LVMH": {"Secteur": "Luxury", "Nom Officiel": "LVMH MOET HENNESSY", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/775670417"},
    "CHRISTIAN DIOR": {"Secteur": "Luxury", "Nom Officiel": "CHRISTIAN DIOR SE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "LOUIS VUITTON": {"Secteur": "Luxury", "Nom Officiel": "LOUIS VUITTON MALLETIER", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "CHRISTIAN LOUBOUTIN": {"Secteur": "Luxury", "Nom Officiel": "CHRISTIAN LOUBOUTIN", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "1 000+ salariés"},
    "CHANEL": {"Secteur": "Luxury", "Nom Officiel": "CHANEL SAS", "Adresse": "Neuilly-sur-Seine (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "HERMES": {"Secteur": "Luxury", "Nom Officiel": "HERMES INTERNATIONAL", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "GUCCI": {"Secteur": "Luxury", "Nom Officiel": "GUCCI", "Adresse": "Florence (Italy)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "PRADA": {"Secteur": "Luxury", "Nom Officiel": "PRADA SPA", "Adresse": "Milan (Italy)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    
    # Transport
    "SNCF": {"Secteur": "Transportation, Logistics & Storage", "Nom Officiel": "SNCF", "Adresse": "Saint-Denis (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/552049447"},
    "RATP": {"Secteur": "Transportation, Logistics & Storage", "Nom Officiel": "REGIE AUTONOME DES TRANSPORTS PARISIENS", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/775663438"},
    "LA POSTE": {"Secteur": "Transportation, Logistics & Storage", "Nom Officiel": "LA POSTE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/356000000"},
    "GROUPE LA POSTE": {"Secteur": "Transportation, Logistics & Storage", "Nom Officiel": "LA POSTE", "Adresse": "Issy-les-Moulineaux (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"}, # Keep existing entry
    "AIR FRANCE": {"Secteur": "Transportation, Logistics & Storage", "Nom Officiel": "AIR FRANCE", "Adresse": "Tremblay-en-France (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/420495178"},

    # Bank / Finance
    "BNP": {"Secteur": "Banking", "Nom Officiel": "BNP PARIBAS", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/662042449"},
    "BNP PARIBAS": {"Secteur": "Banking", "Nom Officiel": "BNP PARIBAS", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/662042449"},
    "SOCIETE GENERALE": {"Secteur": "Banking", "Nom Officiel": "SOCIETE GENERALE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/552120222"},
    "SNOWFLAKE": {"Nom Officiel": "SNOWFLAKE FRANCE", "Secteur": "Tech / Software", "Adresse": "Non renseigné", "Région": "Île-de-France"},
    "ATOS": {"Nom Officiel": "ATOS SE", "Secteur": "Consulting / IT Services", "Adresse": "Bezons (France)", "Région": "Île-de-France", "Effectif": "100 000+ salariés"},
    "CREDIT AGRICOLE": {"Secteur": "Banking", "Nom Officiel": "CREDIT AGRICOLE SA", "Adresse": "Montrouge (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/784608416"},

    # Energy
    "TOTALENERGIES": {"Secteur": "Energy / Utilities", "Nom Officiel": "TOTALENERGIES SE", "Adresse": "Courbevoie (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/542051180"},
    "ENGIE": {"Secteur": "Energy / Utilities", "Nom Officiel": "ENGIE", "Adresse": "Courbevoie (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/542107651"},
    "EDF": {"Secteur": "Energy / Utilities", "Nom Officiel": "ELECTRICITE DE FRANCE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/552081317"},
    
    # Telecom
    "ORANGE": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "ORANGE SA", "Adresse": "Issy-les-Moulineaux (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/380129866"},
    "SFR": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "SFR", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/343059564"},
    "FREE": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "ILIAD (FREE)", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/421938861"},
    "ILIAD": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "ILIAD (FREE)", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"}, # Keep existing entry
    "BOUYGUES TELECOM": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "BOUYGUES TELECOM", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/397480936"},
    
    # Consulting (Added due to user feedback)
    "CAPGEMINI": {"Secteur": "Consulting / IT Services", "Nom Officiel": "CAPGEMINI SE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/330703844"},
    "KPMG": {"Secteur": "Consulting / IT Services", "Nom Officiel": "KPMG S.A", "Adresse": "Paris La Défense (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "DELOITTE": {"Secteur": "Consulting / IT Services", "Nom Officiel": "DELOITTE SAS", "Adresse": "Paris La Défense (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "EY": {"Secteur": "Consulting / IT Services", "Nom Officiel": "ERNST & YOUNG", "Adresse": "Paris La Défense (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "PWC": {"Secteur": "Consulting / IT Services", "Nom Officiel": "PWC FRANCE", "Adresse": "Neuilly-sur-Seine (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "ACCENTURE": {"Secteur": "Consulting / IT Services", "Nom Officiel": "ACCENTURE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    
    # Retail / Grands Magasins
    "GALERIES LAFAYETTE": {"Secteur": "Retail", "Nom Officiel": "GALERIES LAFAYETTE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "PRINTEMPS": {"Secteur": "Retail", "Nom Officiel": "PRINTEMPS", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},

    # Tech / Web
    "SPOTIFY": {"Secteur": "Tech / Software", "Nom Officiel": "SPOTIFY TECHNOLOGY", "Adresse": "Stockholm (Sweden)", "Région": "Monde", "Effectif": "5 000+ salariés"},
    "UBER": {"Secteur": "Tech / Software", "Nom Officiel": "UBER TECHNOLOGIES", "Adresse": "San Francisco, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "AIRBNB": {"Secteur": "Tech / Software", "Nom Officiel": "AIRBNB INC.", "Adresse": "San Francisco, CA (USA)", "Région": "Monde", "Effectif": "5 000+ salariés"},
    "AIR BNB": {"Secteur": "Tech / Software", "Nom Officiel": "AIRBNB INC.", "Adresse": "San Francisco, CA (USA)", "Région": "Monde", "Effectif": "5 000+ salariés"},
    "NETFLIX": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "NETFLIX INC.", "Adresse": "Los Gatos, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "NVIDIA": {"Secteur": "Tech / Software", "Nom Officiel": "NVIDIA CORP", "Adresse": "Santa Clara, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    
    # Auto
    "BMW": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "BMW AG", "Adresse": "Munich (Germany)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "MERCEDES": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "MERCEDES-BENZ GROUP", "Adresse": "Stuttgart (Germany)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "TOYOTA": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "TOYOTA MOTOR CORP", "Adresse": "Toyota City (Japan)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "VOLKSWAGEN": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "VOLKSWAGEN AG", "Adresse": "Wolfsburg (Germany)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    
    # Consumer
    "COCA COLA": {"Secteur": "Food / Beverages", "Nom Officiel": "THE COCA-COLA COMPANY", "Adresse": "Atlanta, GA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés", "Siren": None},
    "DANONE": {"Secteur": "Food / Beverages", "Nom Officiel": "DANONE", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Siren": "552032534"},
    "PEPSI": {"Secteur": "Food / Beverages", "Nom Officiel": "PEPSICO INC.", "Adresse": "Harrison, NY (USA)", "Région": "Monde", "Effectif": "10 000+ salariés", "Siren": None},
    "SAMSUNG": {"Secteur": "Tech / Software", "Nom Officiel": "SAMSUNG ELECTRONICS", "Adresse": "Suwon (South Korea)", "Région": "Monde", "Effectif": "10 000+ salariés", "Siren": None},
    "NIKE": {"Secteur": "Retail", "Nom Officiel": "NIKE INC.", "Adresse": "Beaverton, OR (USA)", "Région": "Monde", "Effectif": "10 000+ salariés", "Siren": None},
    
    # Smartphones / Tech Asia
    "XIAOMI": {"Secteur": "Tech / Software", "Nom Officiel": "XIAOMI CORP", "Adresse": "Beijing (China)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "OPPO": {"Secteur": "Tech / Software", "Nom Officiel": "OPPO ELECTRONICS", "Adresse": "Dongguan (China)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "HUAWEI": {"Secteur": "Tech / Software", "Nom Officiel": "HUAWEI TECHNOLOGIES", "Adresse": "Shenzhen (China)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "ONEPLUS": {"Secteur": "Tech / Software", "Nom Officiel": "ONEPLUS TECHNOLOGY", "Adresse": "Shenzhen (China)", "Région": "Monde", "Effectif": "5 000+ salariés"},

    # Retail / Supermarkets (France & Global)
    "CARREFOUR": {"Secteur": "Retail", "Nom Officiel": "CARREFOUR SA", "Adresse": "Massy (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/652014051"},
    "AUCHAN": {"Secteur": "Retail", "Nom Officiel": "AUCHAN RETAIL", "Adresse": "Croix (France)", "Région": "Hauts-de-France", "Effectif": "10 000+ salariés"},
    "LECLERC": {"Secteur": "Retail", "Nom Officiel": "E.LECLERC", "Adresse": "Ivry-sur-Seine (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "INTERMARCHE": {"Secteur": "Retail", "Nom Officiel": "ITM ENTREPRISES", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "INTERMARCHÉ": {"Secteur": "Retail", "Nom Officiel": "ITM ENTREPRISES", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "LIDL": {"Secteur": "Retail", "Nom Officiel": "LIDL STIFTUNG", "Adresse": "Neckarsulm (Germany)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "ALDI": {"Secteur": "Retail", "Nom Officiel": "ALDI EINKAUF", "Adresse": "Essen (Germany)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "NETTO": {"Secteur": "Retail", "Nom Officiel": "NETTO MARKEN-DISCOUNT", "Adresse": "Germany", "Région": "Monde", "Effectif": "5 000+ salariés"},
    "ACTION": {"Secteur": "Retail", "Nom Officiel": "ACTION B.V.", "Adresse": "Zwaagdijk (Netherlands)", "Région": "Monde", "Effectif": "10 000+ salariés"},

    # Fixes from User Feedback
    "DISNEY": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "THE WALT DISNEY COMPANY", "Adresse": "Burbank, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "DECATHLON": {"Secteur": "Retail", "Nom Officiel": "DECATHLON SE", "Adresse": "Villeneuve-d'Ascq (France)", "Région": "Hauts-de-France", "Effectif": "10 000+ salariés", "Lien": "https://annuaire-entreprises.data.gouv.fr/entreprise/306138900"},
    "LONGCHAMP": {"Secteur": "Luxury", "Nom Officiel": "LONGCHAMP SAS", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "1 000+ salariés"},
    "MONOPRIX": {"Secteur": "Retail", "Nom Officiel": "MONOPRIX", "Adresse": "Clichy (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "PERNOD RICARD": {"Secteur": "Food / Beverages", "Nom Officiel": "PERNOD RICARD", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "PFIZER": {"Secteur": "Pharmaceutics", "Nom Officiel": "PFIZER INC.", "Adresse": "New York, NY (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "LA POSTE": {"Secteur": "Transportation, Logistics & Storage", "Nom Officiel": "LA POSTE", "Adresse": "Issy-les-Moulineaux (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "GROUPE LA POSTE": {"Secteur": "Transportation, Logistics & Storage", "Nom Officiel": "LA POSTE", "Adresse": "Issy-les-Moulineaux (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "TDF": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "TDF", "Adresse": "Montrouge (France)", "Région": "Île-de-France", "Effectif": "1 000+ salariés"},
    "SYMBIO": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "SYMBIO", "Adresse": "Vénissieux (France)", "Région": "Auvergne-Rhône-Alpes", "Effectif": "500+ salariés"},
    "APRIL": {"Secteur": "Insurance / Mutual Health Insurance", "Nom Officiel": "APRIL", "Adresse": "Lyon (France)", "Région": "Auvergne-Rhône-Alpes", "Effectif": "1 000+ salariés"},
    "SAFRAN": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "SAFRAN SA", "Adresse": "Paris (France)", "Région": "Île-de-France", "Effectif": "10 000+ salariés"},
    "VISIATIV": {"Secteur": "Tech / Software", "Nom Officiel": "VISIATIV", "Adresse": "Charbonnières-les-Bains (France)", "Région": "Auvergne-Rhône-Alpes", "Effectif": "1 000+ salariés"},
    "AMOOBI": {"Secteur": "Tech / Software", "Nom Officiel": "AMOOBI", "Adresse": "N/A (International)", "Région": "Monde", "Effectif": "10-50 salariés"},
    "SAFRAN AERO BOOSTERS": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "SAFRAN AERO BOOSTERS", "Adresse": "Herstal (Belgium)", "Région": "Monde", "Effectif": "1 000+ salariés"},
    "SAFRAN AERO BOSOTERS": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "SAFRAN AERO BOOSTERS", "Adresse": "Herstal (Belgium)", "Région": "Monde", "Effectif": "1 000+ salariés"},
    "SERFIGROUP": {"Secteur": "Retail", "Nom Officiel": "SERFI INTERNATIONAL", "Adresse": "Nice (France)", "Région": "Provence-Alpes-Côte d'Azur", "Effectif": "20-49 salariés"},
    "SERFI GROUP": {"Secteur": "Retail", "Nom Officiel": "SERFI INTERNATIONAL", "Adresse": "Nice (France)", "Région": "Provence-Alpes-Côte d'Azur", "Effectif": "20-49 salariés"},
    "SERFI INTERNATIONAL": {"Secteur": "Retail", "Nom Officiel": "SERFI INTERNATIONAL", "Adresse": "Nice (France)", "Région": "Provence-Alpes-Côte d'Azur", "Effectif": "20-49 salariés"},
    
    # International Tech / Electronics (Added for safety)
    "ADOBE": {"Secteur": "Tech / Software", "Nom Officiel": "ADOBE INC.", "Adresse": "San Jose, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "NINTENDO": {"Secteur": "Communication / Media & Entertainment / Telecom", "Nom Officiel": "NINTENDO CO., LTD", "Adresse": "Kyoto (Japan)", "Région": "Monde", "Effectif": "5 000+ salariés"},
    "PHILIPS": {"Secteur": "Manufacturing / Industry", "Nom Officiel": "KONINKLIJKE PHILIPS", "Adresse": "Amsterdam (Netherlands)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    "SALESFORCE": {"Secteur": "Tech / Software", "Nom Officiel": "SALESFORCE", "Adresse": "San Francisco, CA (USA)", "Région": "Monde", "Effectif": "10 000+ salariés"},
    
    # Modern Tech / Remote Tools
    "ZOOM": {"Secteur": "Tech / Software", "Nom Officiel": "ZOOM VIDEO COMMUNICATIONS", "Adresse": "San Jose, CA (USA)", "Région": "Monde", "Effectif": "5 000+ salariés"},
    "SLACK": {"Secteur": "Tech / Software", "Nom Officiel": "SALESFORCE (SLACK)", "Adresse": "San Francisco, CA (USA)", "Région": "Monde", "Effectif": "1 000+ salariés"}
}

# Pre-compute Normalized Keys for Robust Matching
# Maps "COCACOLA" -> "COCA COLA", "LVMH" -> "LVMH", "AIRBNB" -> "AIRBNB"
NORMALIZED_OVERRIDES = {k.replace(" ", "").replace(".", "").replace("-", ""): k for k in GLOBAL_OVERRIDES}

# Explicit Typos / Variations Mapping (Normalized -> Real Key)
NORMALIZED_OVERRIDES["BNBPARIBAS"] = "BNP PARIBAS"
NORMALIZED_OVERRIDES["BNB"] = "BNP PARIBAS"
NORMALIZED_OVERRIDES["BNPPARIBAS"] = "BNP PARIBAS"
NORMALIZED_OVERRIDES["BNP-PARIBAS"] = "BNP PARIBAS"
NORMALIZED_OVERRIDES["FREEPRO"] = "FREE"
NORMALIZED_OVERRIDES["GROUPAGRICA"] = "GROUPE AGRICA"
NORMALIZED_OVERRIDES["MANOMANO"] = "COLIBRI SAS"
NORMALIZED_OVERRIDES["COLIBRI"] = "COLIBRI SAS"
NORMALIZED_OVERRIDES["NATIXIS-CORPORATE-INVESTMENT-BANKING"] = "NATIXIS"
NORMALIZED_OVERRIDES["SOPRA"] = "SOPRA STERIA"
NORMALIZED_OVERRIDES["CLUBMED"] = "CLUB MED"
NORMALIZED_OVERRIDES["CLUB.MED"] = "CLUB MED"
NORMALIZED_OVERRIDES["CLUB-MED"] = "CLUB MED"

def get_region_from_dept(zip_code):
    if not zip_code or len(zip_code) < 2: return "Autre"
    
    # Handle DOM-TOM (3 digits) vs metro (2 digits)
    if zip_code.startswith('97') or zip_code.startswith('98'):
        dept = zip_code[:3]
    else:
        dept = zip_code[:2]

    return DEPT_TO_REGION.get(dept, f"France ({dept})")

def check_is_competitor(name):
    """
    Checks if a company name is a competitor using strict word boundaries.
    Avoids 'EY' matching inside 'DISNEY' or 'KEYRUS'.
    """
    if not name: return False
    upper_name = name.upper()
    for comp in COMPETITORS:
        # \b matches word boundary. escape(comp) handles special chars like &
        if re.search(r'\b' + re.escape(comp) + r'\b', upper_name):
            return True
    return False

# --- Helper Functions ---

def extract_company_from_input(input_str):
    input_str = input_str.strip()
    
    # Pre-cleaning: Text often comes from Excel copy-paste (Tab delimited)
    # Strategy: If tab present, look for the most "name-like" part.
    if "\t" in input_str:
        parts = input_str.split("\t")
        # Heuristic: If part 0 is email, take part 1.
        if "@" in parts[0] and len(parts) > 1:
             input_str = parts[1]
        else:
             input_str = parts[0]

    if "\n" in input_str:
        input_str = input_str.split("\n")[0]
        
    company = input_str.strip()

    # If it's still an email, try to extract domain
    if "@" in company and not company.startswith("http"):
        try:
            domain = company.split("@")[1]
            if "." in domain:
                candidate = domain.split(".")[0]
                # Smart Filter: Keep Gmail/Outlook ignored
                if candidate.lower() not in ["gmail", "outlook", "hotmail", "yahoo", "wanadoo", "icloud", "laposte"]:
                     company = candidate
        except:
            pass
    
    # Heuristics for "Copy-Paste" from directories (Pappers, Societe.com, etc.)
    # Example: "TRANSAVIA a été créée le 1 janvier 1979..." -> "TRANSAVIA"
    # Example: "BNP PARIBAS est une société anonyme..." -> "BNP PARIBAS"
    
    # Regex 1: "X a été créée le"
    match_creation = re.search(r'^(.+?)\s+a été créée le', company, re.IGNORECASE)
    if match_creation:
        company = match_creation.group(1)
        
    # Regex 2: "X est une (société|entreprise|association)"
    if not match_creation:
        match_est = re.search(r'^(.+?)\s+est une\s+(société|entreprise|association)', company, re.IGNORECASE)
        if match_est:
            company = match_est.group(1)

    company = company.replace("-", " ").replace(".", " ")
    company = re.sub(r'(group|france|partners|holdings|corp|inc|ltd)$', r' \1', company, flags=re.IGNORECASE)
    
    return company.strip(), True

def get_sector_from_naf(naf_code):
    if not naf_code: return None
    naf_clean = naf_code.replace(".", "")
    best_sector = None
    max_prefix_len = 0
    for sector, config in SECTOR_CONFIG.items():
        for prefix in config["naf_prefixes"]:
            if naf_clean.startswith(prefix):
                if len(prefix) > max_prefix_len:
                    max_prefix_len = len(prefix)
                    best_sector = sector
    return best_sector

def score_text(text, weights=1.0):
    scores = {}
    text = text.lower()
    for sector, config in SECTOR_CONFIG.items():
        score = 0
        for keyword in config["keywords"]:
            # Valid Regex: r'\b' (Word Boundary). escaped keyword.
            count = len(re.findall(r'\b' + re.escape(keyword) + r'\b', text))
            score += count * weights
        scores[sector] = score
    return scores

def analyze_web_content(company_name):
    try:
        search_results = []
        snippet_text = ""
        source_url = ""
        page_title = "" 
        
        # DuckDuckGo Search (Defensive)
        try:
            # Local import to prevent module-level crash if library is missing/incompatible
            from duckduckgo_search import DDGS
            query = f"{company_name} societe.com France"
            with DDGS() as ddgs:
                 # limit=1
                 results = list(ddgs.text(query, region='fr-fr', max_results=1))
                 if results:
                      first_res = results[0]
                      source_url = first_res.get('href', '')
                      page_title = first_res.get('title', '')
                      snippet_text = f"{page_title} {first_res.get('body', '')}"
        except Exception as e:
            print(f"DDG Lib Error: {e}")
            

        if not snippet_text:
            return None, "URL not found", 0, ""

        # Score the Snippet directly
        # Fix: Use r'\b' for word boundary instead of r'\\b' (which matches literal backslash)
        scores_snippet = score_text(snippet_text, weights=5.0)
        
        final_scores = scores_snippet
            
        if not final_scores or all(score == 0 for score in final_scores.values()):
             return "Unknown", f"Web Analysis ({source_url}) - No keywords in snippet", 0, page_title
             
        best_sector = max(final_scores, key=final_scores.get)
        max_score = final_scores[best_sector]
        
        if max_score > 0:
             return best_sector, f"Web Analysis ({source_url})", max_score, page_title
        
        return "Unknown", f"Web Analysis ({source_url}) - No keywords matched", 0, page_title

    except Exception as e:
        return None, f"Error (Web): {str(e)}", 0, ""


def categorize_company_logic(raw_input):
    try:
        company_name, is_valid = extract_company_from_input(raw_input)
        if not is_valid:
            return {"Input": raw_input, "Nom Officiel": "Ignoré", "Secteur": "Hors Scope", "Détail": "Email perso / invalide", "Source": "-", "Score": "0", "Adresse": "-", "Région": "-", "Lien": "-"}

        # 0. Check User Corrections (Case Insensitive)
        # Force Reload to ensure multi-process / overlapping writes are caught
        load_corrections()
        
        # Key in JSON is UPPERCASE.
        upper_name_clean = normalize_key(company_name)
        
        custom_sector = USER_CORRECTIONS.get(upper_name_clean)
        forced_sector = None
        
        if custom_sector:
            forced_sector = custom_sector
        
        # 1. Check Global Overrides
        target_override = None
        # Try finding override by clean uppercase name
        if upper_name_clean in NORMALIZED_OVERRIDES:
             mapped_key = NORMALIZED_OVERRIDES[upper_name_clean]
             target_override = GLOBAL_OVERRIDES.get(mapped_key)
             # KEY FIX: If we have a better name (Normalized), use it for API search!
             # This helps "groupagrica" -> "GROUPE AGRICA" find results even if no hardcoded override exists.
             company_name = mapped_key # Update for API search
        elif upper_name_clean in GLOBAL_OVERRIDES:
             target_override = GLOBAL_OVERRIDES[upper_name_clean]
        
        # If Override provides explicit address, RETURN IMMEDIATELY (Skip API)
        if target_override and target_override.get("Adresse"):
             manual_link = target_override.get("Lien") 
             if not manual_link:
                  manual_link = f"https://annuaire-entreprises.data.gouv.fr/rechercher?q={target_override['Nom Officiel'].replace(' ', '+')}"
             
             final_sect = forced_sector if forced_sector else target_override["Secteur"]
             
             return {
                "Input": raw_input,
                "Nom Officiel": target_override["Nom Officiel"],
                "Secteur": final_sect,
                "Détail": "Override Global (Hardcoded)",
                "Source": "Base Interne",
                "Score": "100%",
                "Adresse": target_override["Adresse"],
                "Région": target_override["Région"],
                "Effectif": target_override.get("Effectif", "Non renseigné"),
                "Lien": manual_link,
                "IsCompetitor": target_override.get("IsCompetitor", check_is_competitor(target_override["Nom Officiel"]))
             }

        # 2. Call API
        api_url = f"https://recherche-entreprises.api.gouv.fr/search?q={company_name}&per_page=5"
        
        naf_code = None
        official_name = company_name
        address = "Non renseigné"
        region = "Non renseigné"
        link_url = ""
        
        search_success = False

        try:
            response = requests.get(api_url)
            if response.status_code == 200:
                data = response.json()
                if data and data['results']:
                    # Loop to find the first non-CSE/COMITE result
                    best_res = None
                    for res in data['results']:
                         name_check = res.get('nom_complet', '').upper()
                         if "COMITE" not in name_check and "CSE " not in name_check:
                              best_res = res
                              break
                    
                    if not best_res: best_res = data['results'][0]
                    
                    if best_res:
                        search_success = True
                        naf_code = best_res.get('activite_principale')
                        official_name = best_res.get('nom_complet')
                        
                        siege = best_res.get('siege', {})
                        address = siege.get('adresse', best_res.get('adresse', ''))
                        region = siege.get('libelle_region', '')
                        if not region: region = best_res.get('region', '')
                        
                        # Fallback Region from Dept
                        cp = siege.get('code_postal', '')
                        if not region and cp:
                             region = get_region_from_dept(cp)
                        
                        siren = best_res.get('siren')
                        if siren: link_url = f"https://annuaire-entreprises.data.gouv.fr/entreprise/{siren}"
                        
                        # Map Effectif Code to Text
                        tranche_code = best_res.get('tranche_effectif_salarie')
                        effectif_text = TRANCHE_EFFECTIFS.get(tranche_code, "Non renseigné")
                        # If unknown code, keep it raw or default
                        if not effectif_text and tranche_code: effectif_text = f"Code: {tranche_code}"
                        best_res['tranche_effectif_salarie'] = effectif_text
                        
        except Exception as e:
            print(f"API Call Error: {e}")

        # 3. Determine Final Result
        if search_success:
            # If we had a forced sector from overrides, use it
            final_sector = forced_sector if forced_sector else get_sector_from_naf(naf_code)
            
            # If still unknown sector, use partial override if exists
            if not final_sector and target_override:
                final_sector = target_override.get("Secteur", "Unknown")
            
            if not final_sector: final_sector = "Unknown"

            # Check Competitor (Strict Word Boundary Match)
            # Check Competitor (Strict Word Boundary Match)
            # Use the new robust helper function
            is_competitor = check_is_competitor(official_name)
            
            # Additional Check: If forced_sector name matches competitor list
            if not is_competitor and forced_sector and forced_sector.upper() in COMPETITORS:
                 # Unlikely case but safety check
                 pass

            return {
                "Input": raw_input,
                "Nom Officiel": official_name,
                "Secteur": final_sector,
                "Secteur": final_sector,
                "Détail": "Override + API" if forced_sector else f"Code NAF: {naf_code}",
                "Source": "Officiel (API)",
                "Score": "100%",
                "Adresse": address,
                "Région": region,
                "Lien": link_url,
                "IsCompetitor": is_competitor,
                "Effectif": best_res.get('tranche_effectif_salarie')
            }
            
        # 4. Fallback: Web Search
        # If API failed, but we have a partial override (without address), usually we returned above?
        # But if we are here, we have neither robust API result nor specific override address.
        # Check overrides one last time for sector only?
        if forced_sector:
             return {
                "Input": raw_input,
                "Nom Officiel": company_name,
                "Secteur": forced_sector,
                "Détail": "Correction Utilisateur (Sans Info)",
                "Source": "Mémoire",
                "Score": "100%",
                "Adresse": "-", "Région": "-", "Lien": "-",
                "Adresse": "-", "Région": "-", "Lien": "-",
                "IsCompetitor": check_is_competitor(company_name)
             }
             
        sector_web, source_web, score_web, title_web = analyze_web_content(company_name)
        
        final_link = link_url
        if final_link == "-" or not final_link:
             final_link = f"https://annuaire-entreprises.data.gouv.fr/rechercher?q={company_name.replace(' ', '+')}"

        # Fix: Only accept Web Result if it is NOT "Unknown"
        if sector_web and sector_web != "Unknown":
             return {
                "Input": raw_input,
                "Nom Officiel": title_web if title_web and len(title_web) < 60 else official_name,
                "Secteur": sector_web,
                "Détail": f"Web Analysis ({score_web})",
                "Source": source_web,
                "Score": f"{score_web}",
                "Adresse": address if address != "Non renseigné" else "International / Web",
                "Région": region if region != "Non renseigné" else "Monde",
                "Lien": final_link,
                "IsCompetitor": check_is_competitor(official_name) or check_is_competitor(company_name)
             }
             
        # 5. Fallback AI (Groq) - Last Resort
        print(f"Triggering Groq for: {company_name}")
        ai_sector, ai_detail, ai_score = analyze_with_groq(company_name, list(SECTOR_CONFIG.keys()), CUSTOM_SECTORS)
        
        if ai_sector:
             return {
                "Input": raw_input,
                "Nom Officiel": official_name, # Keep best guess official name
                "Secteur": ai_sector,
                "Détail": ai_detail,
                "Source": "Intelligence Artificielle (Groq)",
                "Score": "100%",
                "Adresse": address,
                "Région": region,
                "Lien": final_link,
                "IsCompetitor": check_is_competitor(official_name)
             }

        
        # 6. Degraded Mode: AI Failed, but we had a Web Trace
        # If we have a URL/Title from Web Search, use it even if sector keywords were not found.
        # This prevents blocking the user when AI quota is exceeded.
        if sector_web == "Unknown" and title_web and source_web:
             detail_msg = "Mode Dégradé (Web)"
             if 'ai_detail' in locals() and ai_detail:
                 detail_msg = f"Web (AI HS: {ai_detail})"
             
             return {
                "Input": raw_input,
                "Nom Officiel": title_web if len(title_web) < 60 else official_name,
                "Secteur": "À Vérifier / Hors Liste",
                "Détail": detail_msg,
                "Source": source_web,
                "Score": "10% (Web)",
                "Adresse": address if address != "Non renseigné" else "International / Web",
                "Région": region if region != "Non renseigné" else "Monde",
                "Lien": final_link,
                "IsCompetitor": check_is_competitor(title_web) or check_is_competitor(company_name)
             }

        # 7. Nothing Found
        detail_msg = "Aucun résultat probant"
        if 'ai_detail' in locals() and ai_detail:
             detail_msg = f"Echec AI: {ai_detail}"

        return {
            "Input": raw_input,
            "Nom Officiel": official_name,
            "Secteur": "Non Trouvé",
            "Détail": detail_msg,
            "Source": "-",
            "Score": "0",
            "Adresse": "-", "Région": "-", "Lien": "-",
            "IsCompetitor": check_is_competitor(official_name)
        }

    except Exception as e:
        return { 
            "Input": raw_input, 
            "Nom Officiel": "Erreur", 
            "Secteur": "Erreur", 
            "Détail": str(e), 
            "Source": "Crash", 
            "Score": "0", 
            "Adresse": "-", "Région": "-", "Lien": "-" 
        }

# --- Routes ---

@app.route('/')
def index():
    # Combine standard + custom sectors
    all_sectors = sorted(list(SECTOR_CONFIG.keys()) + CUSTOM_SECTORS)
    return render_template('index.html', sectors=all_sectors, custom_sectors=CUSTOM_SECTORS)

@app.route('/api/categorize', methods=['POST'])
def api_categorize():
    data = request.json
    company_input = data.get('input')
    if not company_input:
        return jsonify({"error": "No input provided"}), 400
    
    result = categorize_company_logic(company_input)
    return jsonify(result)

@app.route('/api/override', methods=['POST'])
def override_sector():
    data = request.json
    name = data.get('name')
    sector = data.get('sector')
    
    if name and sector:
        # Save the correction using the EXTRACTED/NORMALIZED name as the key (UPPERCASE)
        # This ensures that future searches (which also use the extracted name) find the correction.
        # Example: Input "dlv@foo.com" -> Extracted "foo". Saving "FOO": "Sector" fixes "contact@foo.com" too.
        normalized_name, _ = extract_company_from_input(name)
        save_correction(normalized_name.upper(), sector)
        
        # Determine if it's a new custom sector
        standard_sectors = list(SECTOR_CONFIG.keys())
        if sector not in standard_sectors and sector not in CUSTOM_SECTORS:
            CUSTOM_SECTORS.append(sector)
            save_custom_sectors()
            
        return jsonify({"status": "success", "sector": sector, "is_new": sector in CUSTOM_SECTORS})
    return jsonify({"error": "Missing data"}), 400

@app.route('/api/delete_sector', methods=['POST'])
def delete_sector():
    data = request.json
    sector = data.get('sector')
    
    if sector in CUSTOM_SECTORS:
        CUSTOM_SECTORS.remove(sector)
        save_custom_sectors()
        return jsonify({"status": "success", "message": "Sector deleted"})
    return jsonify({"error": "Sector not found or cannot delete standard sector"}), 400

@app.route('/api/upload', methods=['POST'])
def api_upload():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    
    filename = file.filename.lower()
    inputs = []
    
    try:
        if filename.endswith('.csv'):
            # Read CSV using standard library
            stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
            csv_input = csv.reader(stream)
            for row in csv_input:
                if row:
                     inputs.append(str(row[0])) # Take first column

        elif filename.endswith(('.xls', '.xlsx')):
            # Read Excel using openpyxl
            wb = load_workbook(file)
            ws = wb.active
            # Iterate rows, take first column
            for row in ws.iter_rows(values_only=True):
                if row and row[0]:
                    inputs.append(str(row[0]))
        else:
             return jsonify({"error": "Format non supporté (CSV ou Excel)"}), 400
             
        # Process the list (same as batch)
        results = []
        for i, line in enumerate(inputs):
            if line.strip():
                 if i > 0: time.sleep(1.0) # Rate limit
                 results.append(categorize_company_logic(line))
        
        return jsonify({"results": results})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/batch', methods=['POST'])
def api_batch():
    try:
        data = request.json
        if not data:
             return jsonify({"error": "Invalid JSON/Empty Body"}), 400
             
        inputs = data.get('inputs', [])
        results = []
        
        for i, line in enumerate(inputs):
            if not isinstance(line, str): line = str(line) # Safety cast
            
            if line.strip():
                # Basic rate limiting for batch to avoid flooding Google if many requests
                if i > 0: time.sleep(1.0)
                try:
                    results.append(categorize_company_logic(line))
                except Exception as e:
                    print(f"Batch Error on {line}: {e}")
        
        return jsonify({"results": results})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    try:
        data = request.json
        if not data or 'results' not in data:
            return jsonify({"error": "No data to export"}), 400
        
        results = data['results']
        
        # Prepare DataFrame
        # Prepare Data for Excel
        headers = ["Input", "Nom Officiel", "Secteur", "Adresse", "Région", "Effectif", "Lien", "Score", "Détails"]
        
        # Save to a temporary file (or in memory)
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultats"
        
        # Write Header
        ws.append(headers)
        
        # Write Data
        for row in results:
            ws.append([
                row.get("Input"),
                row.get("Nom Officiel"),
                row.get("Secteur"),
                row.get("Adresse"),
                row.get("Région"),
                row.get("Effectif"),
                row.get("Lien"),
                row.get("Score"),
                row.get("Détail")
            ])
            
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
        
        # Apply Filter
        ws.auto_filter.ref = ws.dimensions
        
        # Freeze Top Row
        ws.freeze_panes = 'A2'
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Style Body and Adjust Width
        # Calculate max width for each column
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                if len(column_widths) <= i:
                    column_widths.append(0)
                length = len(str(cell.value)) if cell.value else 0
                if length > column_widths[i]:
                    column_widths[i] = length
                    
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[chr(65 + i)].width = min(column_width + 4, 60)
            
        # Borders for all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                     cell.alignment = Alignment(vertical="center")

        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"export_entreprises_{int(time.time())}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
            
    except Exception as e:
        print(f"Export Error: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=True)
